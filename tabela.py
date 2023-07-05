import streamlit as st
import pandas as pd
import base64
import re
import locale
from openpyxl import load_workbook

# Cria a interface Streamlit
st.set_page_config(page_title='Tratador de Tabelas', page_icon='img/icone.ico')

# Importa o arquivo CSS


def local_css(file_name):
    with open(file_name) as f:
        st.markdown(f'<style>{f.read()}</style>', unsafe_allow_html=True)


# Chama a função para aplicar o estilo
local_css('style.css')

# código JavaScript para controlar o clique do botão de download
st.markdown(
    """
    <script>
    const downloadButton = document.getElementById('download-button');
    downloadButton.addEventListener('click', function() {
        downloadButton.classList.add('btn-clicked');
    });
    </script>
    """,
    unsafe_allow_html=True
)

# Define o local como pt-BR para formatar os valores como R$
locale.setlocale(locale.LC_ALL, 'pt_BR')

# Caminho do arquivo CSV da tabela base
caminho_tabela_base = 'teste/base.csv'

# Carrega a tabela base em um DataFrame
df_base = pd.read_csv(caminho_tabela_base, sep=';', quoting=0)

# Função para tratar a tabela


# Função para tratar a tabela
def tratar_tabela(caminho_arquivo):
    global df_final  # Declarando a variável df_final como global

    if caminho_arquivo is None:
        st.warning('Nenhum arquivo selecionado.')
        return

    # Carrega o arquivo Excel em um objeto Workbook
    try:
        wb = load_workbook(caminho_arquivo, read_only=True)
    except Exception as e:
        st.warning(f"Erro ao carregar o arquivo: {str(e)}")
        return

    # Obtém a lista de nomes de abas (sheets) no arquivo
    nomes_abas = wb.sheetnames

    # Verifica se há abas no arquivo
    if len(nomes_abas) == 0:
        st.warning('Nenhuma aba encontrada no arquivo.')
        return

    # Cria uma tabela final vazia
    df_final = pd.DataFrame(columns=['código','Índice', 'Porte', 'ch', 'Filme', 'mnemônico',  'Efetua',  'VlrPorte'])

    # Itera sobre cada aba e processa os dados de cada uma
    for nome_aba in nomes_abas:
        # Carrega a aba em um DataFrame
        df = pd.read_excel(caminho_arquivo, sheet_name=nome_aba)

        # Verifica se o DataFrame está vazio
        if df.empty:
            st.warning(f"A aba '{nome_aba}' está vazia. Será pulada.")
            continue

        # Seleciona as colunas desejadas
        colunas = list(df.columns)
        colunas_selecionadas = st.multiselect(
            f'Selecione as colunas desejadas ({nome_aba})', colunas, format_func=lambda x: x, default=[])

        # Verifica se alguma coluna foi selecionada
        if len(colunas_selecionadas) == 0:
            st.warning('Nenhuma coluna selecionada.')
            return

        # Lista de palavras-chave para identificar as colunas de código, mnemônico e valor
        palavras_chave_codigo = ['codigo', 'código', 'CODIGO']
        palavras_chave_mnemonico = ['mnemonico', 'mnemônico', 'MNEMONICO']
        palavras_chave_valor = ['valor', 'VALOR']

        # Busca a coluna de código
        coluna_codigo = next(
            (col for col in colunas_selecionadas if any(keyword in col.lower() for keyword in palavras_chave_codigo)), None)

        # Busca a coluna de mnemônico
        coluna_mnemonico = next(
            (col for col in colunas_selecionadas if any(keyword in col.lower() for keyword in palavras_chave_mnemonico)), None)

        # Busca a coluna de valor
        coluna_valor = next(
            (col for col in colunas_selecionadas if any(keyword in col.lower() for keyword in palavras_chave_valor)), None)

        # Trata os dados da tabela base de acordo com as colunas selecionadas
        df_base_atualizada = pd.DataFrame()

        if coluna_codigo is not None:
            df_base_atualizada['código'] = df[coluna_codigo].astype(str).str.replace(r'\D+', '', regex=True)
        else:
            df_base_atualizada['código'] = ''

        if coluna_mnemonico is not None:
            if coluna_codigo is None:
                df_base_atualizada['código'] = ''
            df_base_atualizada['mnemônico'] = df[coluna_mnemonico]
        else:
            if coluna_codigo is None:
                df_base_atualizada['código'] = ''

        if coluna_valor is not None:
            df_base_atualizada['ch'] = df[coluna_valor].apply(
                lambda x: format_currency(x) if pd.notnull(x) else '')

        # Renomeia as colunas padrão
        colunas_padrao = ['Índice', 'Porte', 'código', 'mnemônico', 'Filme', 'Efetua', 'ch', 'VlrPorte']
        for col in colunas_padrao:
            if col not in df_base_atualizada.columns:
                df_base_atualizada[col] = None

        # Preenche as colunas padrão com os valores fornecidos
        df_base_atualizada['Índice'] = 0
        df_base_atualizada['Porte'] = 'UNIL'
        df_base_atualizada['Filme'] = 0
        df_base_atualizada['Efetua'] = 'S'
        df_base_atualizada['VlrPorte'] = 1

        


        # Adiciona os dados da aba atual na tabela final
        df_final = pd.concat([df_final, df_base_atualizada[['código', 'Índice', 'Porte', 'ch', 'Filme', 'mnemônico', 'Efetua', 'VlrPorte']]], ignore_index=True)


    # Remove os 0 a mais que possam ser adicionados nos códigos
    if 'código' in df_final.columns:
        df_final['código'] = df_final['código'].astype(str).str.rstrip('0').str.ljust(8, '0').replace('nan00000', '')
    else:
        df_final['código'] = None


    # Remove linhas totalmente vazias da tabela final
    df_final = df_final.dropna(how='all', subset=['código', 'mnemônico', 'ch']).reset_index(drop=True)

    # Mostra um preview do DataFrame tratado
    st.subheader('Tabela Tratada')
    st.dataframe(df_final)

    # Verifica se a tabela final está vazia
    if df_final.empty:
        st.warning('A tabela tratada está vazia.')
        return

    # Codifica o DataFrame tratado em CSV para download
    csv = df_final.to_csv(sep=';', quoting=0, index=False)
    b64 = base64.b64encode(csv.encode()).decode()
    href = f'<a href="data:file/csv;base64,{b64}" download="arquivo_tratado.csv" class="btn-download clicked">Baixar Tabela Tratada</a>'

    st.markdown(href, unsafe_allow_html=True)

# Função para formatar o valor da coluna "ch", tirando as letras e $, e com 2 casas decimais
def format_currency(value):
    if pd.notnull(value):
        value = str(value)
        if value.startswith('RS '):
            value = value[3:]
        value = re.sub(r'[^0-9,.]', '', value)
        try:
            value = locale.currency(float(value), grouping=True, symbol=False)
        except ValueError:
            pass
    return value

# Chama a função para aplicar o estilo
local_css('style.css')

# Define o local como pt-BR para formatar os valores como R$
locale.setlocale(locale.LC_ALL, 'pt_BR')

st.image('img/tabela.png', width=250)
st.info('Antes de selecionar o arquivo, certifique-se de que as colunas dos códigos de procedimento, do valor que você quer usar e do mnemônico (caso tenha) estão com os títulos escritos corretamente, assim: "código" ou "codigo", "valor", "mnemonico" ou "mnemônico"', icon="🚨")

# Solicita o caminho do arquivo Excel
caminho_arquivo = st.file_uploader(
    'Selecione o arquivo Excel', type=['xls', 'xlsx'])

# Trata a tabela
tratar_tabela(caminho_arquivo)
